"""
Moduł tworzący plik Excel z dashboardem na podstawie danych ze scrapera.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from collections import defaultdict
from datetime import datetime


# ─── KOLORY ───────────────────────────────────────────────────────
C_NAVY    = "1B2A4A"
C_BLUE    = "2563EB"
C_BLUE_L  = "DBEAFE"
C_BLUE_LL = "EFF6FF"
C_GREEN   = "059669"
C_GREEN_L = "D1FAE5"
C_AMBER   = "D97706"
C_AMBER_L = "FEF3C7"
C_RED     = "DC2626"
C_RED_L   = "FEE2E2"
C_GRAY    = "6B7280"
C_GRAY_L  = "F3F4F6"
C_WHITE   = "FFFFFF"
C_BORDER  = "E5E7EB"
C_HEADER_BG = "1E3A5F"


def side(color=C_BORDER, style="thin"):
    return Side(border_style=style, color=color)


def border_all(color=C_BORDER):
    s = side(color)
    return Border(left=s, right=s, top=s, bottom=s)


def border_bottom(color=C_BORDER):
    return Border(bottom=side(color))


def fill(color):
    return PatternFill("solid", fgColor=color)


def font(bold=False, color=C_NAVY, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)


def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ─── POMOCNICZE ───────────────────────────────────────────────────

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def header_cell(ws, row, col, value, bg=C_HEADER_BG, fg=C_WHITE, size=10, center=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font(bold=True, color=fg, size=size)
    cell.fill = fill(bg)
    cell.alignment = align("center" if center else "left")
    cell.border = border_all(C_HEADER_BG)
    return cell


def metric_block(ws, row, col, label, value, sub=None,
                 bg=C_BLUE_LL, val_color=C_BLUE):
    """Rysuje kafelek KPI: label / value / sub."""
    # Merge
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + 2)
    ws.merge_cells(start_row=row + 1, start_column=col,
                   end_row=row + 1, end_column=col + 2)
    if sub:
        ws.merge_cells(start_row=row + 2, start_column=col,
                       end_row=row + 2, end_column=col + 2)

    lbl = ws.cell(row=row, column=col, value=label)
    lbl.font = font(size=9, color=C_GRAY)
    lbl.fill = fill(bg)
    lbl.alignment = align("center")

    val = ws.cell(row=row + 1, column=col, value=value)
    val.font = font(bold=True, size=16, color=val_color)
    val.fill = fill(bg)
    val.alignment = align("center")

    if sub:
        s = ws.cell(row=row + 2, column=col, value=sub)
        s.font = font(size=8, color=C_GRAY, italic=True)
        s.fill = fill(bg)
        s.alignment = align("center")

    # obramowanie bloku
    for r in range(row, row + (3 if sub else 2)):
        for c in range(col, col + 3):
            ws.cell(row=r, column=c).border = border_all(C_BLUE_L)


# ─── ARKUSZE ──────────────────────────────────────────────────────

def build_dashboard(ws, oferty):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_BLUE

    # Kolumny: A(margines) B-F(tabela lewa) G(odstęp) H-N(wykres prawy)
    # Szerokości: A=1.5, B=18, C=10, D=14, E=14, F=14, G=2, H-N=8 każda
    col_widths = [1.5, 18, 10, 14, 14, 14, 2, 8, 8, 8, 8, 8, 8, 8]
    for i, w in enumerate(col_widths, 1):
        set_col_width(ws, i, w)

    for r in range(1, 80):
        ws.row_dimensions[r].height = 18

    # ── NAGŁÓWEK ────────────────────────────────────────────────
    ws.merge_cells("B1:N1")
    t = ws["B1"]
    t.value = "RAPORT RYNKU NIERUCHOMOŚCI"
    t.font = Font(name="Arial", bold=True, size=20, color=C_WHITE)
    t.fill = fill(C_NAVY)
    t.alignment = align("left", "center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("B2:N2")
    sub = ws["B2"]
    sub.value = f"Analiza {len(oferty)} ofert · wygenerowano {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    sub.font = font(size=9, color="94A3B8")
    sub.fill = fill(C_NAVY)
    sub.alignment = align("left", "center")

    # Odstęp po nagłówku
    ws.merge_cells("B3:N3")
    ws["B3"].fill = fill("F8FAFF")
    ws.row_dimensions[3].height = 8

    # ── KPI — dynamiczne wg typu transakcji ─────────────────────
    sprzedaz = [o for o in oferty if o.transakcja == "Sprzedaż"]
    wynajem  = [o for o in oferty if o.transakcja == "Wynajem"]
    jest_sprzedaz = len(sprzedaz) > 0
    jest_wynajem  = len(wynajem) > 0

    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 8

    # Kafelek 1 — zawsze: łączna liczba ofert
    metric_block(ws, 4, 2, "Liczba ofert", len(oferty),
                 None, C_BLUE_LL, C_BLUE)

    if jest_sprzedaz and jest_wynajem:
        # Rzadki przypadek: mieszane dane — pokaż oba
        avg_m2  = sum(o.cena_za_m2 for o in sprzedaz) / len(sprzedaz)
        avg_wyn = sum(o.cena for o in wynajem) / len(wynajem)
        metric_block(ws, 4, 6, "Średnia cena/m² (sprzedaż)",
                     f"{avg_m2:,.0f} zł", f"{len(sprzedaz)} ofert", C_GREEN_L, C_GREEN)
        metric_block(ws, 4, 10, "Średni czynsz/m² (czynsz)",
                     f"{avg_wyn:,.0f} zł/mies.", f"{len(wynajem)} ofert", C_AMBER_L, C_AMBER)
    elif jest_sprzedaz:
        avg_m2   = sum(o.cena_za_m2 for o in sprzedaz) / len(sprzedaz)
        avg_cena = sum(o.cena for o in sprzedaz) / len(sprzedaz)
        metric_block(ws, 4, 6, "Średnia cena/m² (sprzedaż)",
                     f"{avg_m2:,.0f} zł", f"{len(sprzedaz)} ofert", C_GREEN_L, C_GREEN)
        metric_block(ws, 4, 10, "Średnia cena",
                     f"{avg_cena:,.0f} zł", "wszystkie miasta", C_AMBER_L, C_AMBER)
    else:
        avg_wyn = sum(o.cena for o in wynajem) / len(wynajem) if wynajem else 0
        avg_m2  = sum(o.cena_za_m2 for o in wynajem) / len(wynajem) if wynajem else 0
        metric_block(ws, 4, 6, "Śr. czynsz miesięczny",
                     f"{avg_wyn:,.0f} zł/mies.", f"{len(wynajem)} ofert wynajmu", C_AMBER_L, C_AMBER)
        metric_block(ws, 4, 10, "Śr. czynsz/m²",
                     f"{avg_m2:,.0f} zł/m²", "wszystkie miasta", C_GREEN_L, C_GREEN)

    # ── SEKCJA LEWA: tabela wg miasta ───────────────────────────
    # Dane do tabeli i wykresu — bierzemy dostępną grupę
    grupa_glowna = sprzedaz if jest_sprzedaz else wynajem
    tytul_tabeli = (
        "Średnia cena/m² wg miasta" if jest_sprzedaz
        else "Średni czynsz wg miasta"
    )
    etykieta_osi = "zł/m²" if jest_sprzedaz else "zł/mies."
    format_komorki = '#,##0 "zł"'

    ROW_TABELA = 9   # stały wiersz startowy tabeli — nie zależy od KPI

    ws.merge_cells(f"B{ROW_TABELA}:F{ROW_TABELA}")
    th = ws[f"B{ROW_TABELA}"]
    th.value = tytul_tabeli
    th.font = font(bold=True, size=11, color=C_NAVY)
    th.alignment = align("left")
    ws.row_dimensions[ROW_TABELA].height = 24

    row = ROW_TABELA + 1
    cities_data = defaultdict(list)
    for o in grupa_glowna:
        val = o.cena_za_m2 if jest_sprzedaz else o.cena
        cities_data[o.miasto].append(val)

    headers_tab = ["Miasto", "Ofert", "Średnia", "Min", "Max"]
    for ci, h in enumerate(headers_tab):
        header_cell(ws, row, ci + 2, h)

    chart_data_start = row + 1
    for ri, (miasto, wartosci) in enumerate(
        sorted(cities_data.items(), key=lambda x: -sum(x[1]) / len(x[1]))
    ):
        r2 = row + 1 + ri
        bg = C_GRAY_L if ri % 2 == 0 else C_WHITE
        avg_v = sum(wartosci) / len(wartosci)
        row_vals = [miasto, len(wartosci), avg_v, min(wartosci), max(wartosci)]
        for ci, v in enumerate(row_vals):
            c = ws.cell(row=r2, column=ci + 2,
                        value=round(v) if isinstance(v, float) else v)
            c.font = font(color=C_NAVY if ci == 0 else C_GRAY,
                          bold=(ci == 0), size=9)
            c.fill = fill(bg)
            c.alignment = align("right" if ci > 0 else "left")
            c.border = border_all()
            if ci in (2, 3, 4):
                c.number_format = format_komorki
        ws.row_dimensions[r2].height = 17

    chart_data_end = row + max(len(cities_data), 1)

    # ── WYKRES 1: słupkowy (kolumna H, wiersz 9) ────────────────
    # Umieszczamy go na prawo od tabeli, zaczynając od H9.
    # Wysokość: 14 wierszy (ok. 252px przy 18px/wiersz)
    bc = BarChart()
    bc.type = "bar"
    bc.title = None
    bc.style = 10
    bc.grouping = "clustered"
    bc.y_axis.numFmt = '#,##0'
    bc.y_axis.title = etykieta_osi
    bc.x_axis.title = None
    bc.legend = None
    bc.width  = 14   # cm
    bc.height = 10   # cm

    if len(cities_data) >= 1:
        data_ref = Reference(ws, min_col=4, min_row=chart_data_start,
                             max_row=chart_data_end)
        cats_ref = Reference(ws, min_col=2, min_row=chart_data_start,
                             max_row=chart_data_end)
        bc.add_data(data_ref)
        bc.set_categories(cats_ref)
        bc.series[0].graphicalProperties.solidFill = C_BLUE

    WYKRES1_KOMORKA = "H9"    # wykres słupkowy zawsze tu
    ws.add_chart(bc, WYKRES1_KOMORKA)

    # ── SEKCJA LEWA: tabela wg typu nieruchomości ───────────────
    # Zaczyna się BEZPOŚREDNIO pod tabelą miast, z 2-wierszowym odstępem
    row_typ = chart_data_end + 2

    ws.merge_cells(f"B{row_typ}:F{row_typ}")
    th2 = ws[f"B{row_typ}"]
    th2.value = "Rozkład ofert wg typu nieruchomości"
    th2.font = font(bold=True, size=11, color=C_NAVY)
    th2.alignment = align("left")
    ws.row_dimensions[row_typ].height = 24

    row_typ += 1
    type_counts = defaultdict(int)
    for o in oferty:
        type_counts[o.typ] += 1

    header_cell(ws, row_typ, 2, "Typ")
    header_cell(ws, row_typ, 3, "Liczba")
    header_cell(ws, row_typ, 4, "Udział %")

    pie_start = row_typ + 1
    for ri, (typ, cnt) in enumerate(sorted(type_counts.items(), key=lambda x: -x[1])):
        r2 = row_typ + 1 + ri
        bg = C_GRAY_L if ri % 2 == 0 else C_WHITE
        for ci, v in enumerate([typ, cnt, cnt / len(oferty)]):
            c = ws.cell(row=r2, column=ci + 2, value=v)
            c.font = font(size=9)
            c.fill = fill(bg)
            c.alignment = align("right" if ci > 0 else "left")
            c.border = border_all()
            if ci == 2:
                c.number_format = "0.0%"
        ws.row_dimensions[r2].height = 17
    pie_end = row_typ + len(type_counts)

    # ── WYKRES 2: kołowy — umieszczamy POD wykresem słupkowym ───
    # Wykres słupkowy ma height=10cm ≈ 14 wierszy przy 18px.
    # Zaczynamy wykres kołowy od wiersza 9+14 = 23 (z małym zapasem: 24).
    WYKRES2_ROW = 24
    WYKRES2_KOMORKA = f"H{WYKRES2_ROW}"

    pc = PieChart()
    pc.title = None
    pc.style = 10
    pc.width  = 14   # cm — taka sama szerokość jak słupkowy
    pc.height = 10   # cm

    pdata = Reference(ws, min_col=3, min_row=pie_start, max_row=pie_end)
    pcats = Reference(ws, min_col=2, min_row=pie_start, max_row=pie_end)
    pc.add_data(pdata)
    pc.set_categories(pcats)
    ws.add_chart(pc, WYKRES2_KOMORKA)

    # ── STOPKA ───────────────────────────────────────────────────
    footer_row = max(pie_end, WYKRES2_ROW + 14) + 3
    ws.merge_cells(f"B{footer_row}:N{footer_row}")
    f_cell = ws[f"B{footer_row}"]
    f_cell.value = (
        "damian.bisewski@gmail.com"
    )
    f_cell.font = font(size=8, color=C_GRAY, italic=True)
    f_cell.alignment = align("center")


def build_dane(ws, oferty):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_NAVY

    columns = [
        ("ID", 8),
        ("Tytuł", 40),
        ("Miasto", 12),
        ("Dzielnica", 16),
        ("Typ", 12),
        ("Transakcja", 12),
        ("Cena (zł)", 14),
        ("Powierzchnia (m²)", 16),
        ("Pokoje", 8),
        ("Piętro", 8),
        ("Rok budowy", 12),
        ("Cena/m² (zł)", 14),
        ("Data dodania", 13),
        ("URL", 40),
    ]

    for ci, (name, width) in enumerate(columns, 1):
        header_cell(ws, 1, ci, name)
        set_col_width(ws, ci, width)

    ws.row_dimensions[1].height = 22

    for ri, o in enumerate(oferty, 2):
        bg = C_GRAY_L if ri % 2 == 0 else C_WHITE
        values = [
            o.id, o.tytul, o.miasto, o.dzielnica, o.typ, o.transakcja,
            o.cena, o.powierzchnia, o.liczba_pokoi,
            o.pietro if o.pietro is not None else "—",
            o.rok_budowy, o.cena_za_m2, o.data_dodania, o.url,
        ]
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = font(size=9)
            c.fill = fill(bg)
            c.alignment = align("left")
            c.border = border_all()
            if ci in (7, 12):
                c.number_format = '#,##0 "zł"'
            if ci == 8:
                c.number_format = '0.0 "m²"'
        ws.row_dimensions[ri].height = 16

    # Conditional formatting: cena/m²
    last_row = len(oferty) + 1
    ws.conditional_formatting.add(
        f"L2:L{last_row}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        )
    )

    # Tabela — tylko gdy jest przynajmniej 1 wiersz danych
    if last_row >= 2:
        table = Table(displayName="Nieruchomosci", ref=f"A1:N{last_row}")
        style = TableStyleInfo(name="TableStyleMedium2",
                               showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{last_row}"


def build_analiza(ws, oferty):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_GREEN

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "ANALIZA RYNKU — TOP OFERTY I STATYSTYKI"
    t.font = font(bold=True, size=14, color=C_WHITE)
    t.fill = fill(C_NAVY)
    t.alignment = align("center")
    ws.row_dimensions[1].height = 32

    set_col_width(ws, 1, 40)
    set_col_width(ws, 2, 14)
    set_col_width(ws, 3, 14)
    set_col_width(ws, 4, 14)
    set_col_width(ws, 5, 14)
    set_col_width(ws, 6, 12)

    # Top 10 najtańsze/m²
    row = 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "TOP 10 — Najlepsza cena za m² (sprzedaż)"
    ws[f"A{row}"].font = font(bold=True, size=11, color=C_NAVY)
    ws[f"A{row}"].alignment = align("left")
    ws.row_dimensions[row].height = 22
    row += 1

    for h, c in zip(["Oferta", "Miasto", "Typ", "Pow. m²", "Cena/m²", "Cena total"],
                    range(1, 7)):
        header_cell(ws, row, c, h)
    row += 1

    sprzedaz = sorted(
        [o for o in oferty if o.transakcja == "Sprzedaż"],
        key=lambda x: x.cena_za_m2
    )
    for ri, o in enumerate(sprzedaz[:10]):
        bg = C_GREEN_L if ri % 2 == 0 else C_WHITE
        for ci, v in enumerate([o.url, o.miasto, o.typ,
                                 o.powierzchnia, o.cena_za_m2, o.cena], 1):
            c = ws.cell(row=row + ri, column=ci, value=v)
            c.font = font(size=9)
            c.fill = fill(bg)
            c.border = border_all()
            c.alignment = align("left" if ci == 1 else "right")
            if ci in (5, 6):
                c.number_format = '#,##0 "zł"'
            if ci == 4:
                c.number_format = '0.0 "m²"'
        ws.row_dimensions[row + ri].height = 16

    # Top 10 wynajem
    row += 12
    ws.merge_cells(f"A{row}:F{row}")


# ─── MAIN ─────────────────────────────────────────────────────────

def generuj_excel(oferty, output_path="raport_nieruchomosci.xlsx"):
    print("[INFO] Tworzenie pliku Excel...")
    wb = Workbook()

    ws_dash = wb.active
    ws_dash.title = "📊 Dashboard"

    ws_dane = wb.create_sheet("📋 Dane")
    ws_analiza = wb.create_sheet("🏆 Analiza")

    build_dashboard(ws_dash, oferty)
    build_dane(ws_dane, oferty)
    build_analiza(ws_analiza, oferty)

    wb.save(output_path)
    print(f"[OK]   Zapisano: {output_path}\n")
    return output_path
